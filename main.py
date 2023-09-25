import shutil
from os import listdir, getcwd, path

import openpyxl as op
from openpyxl import Workbook
from progress.bar import IncrementalBar
from tqdm import trange


def read_cicle(filename: str, solder: bool) -> list:
    """
    Функция формирует список словарей с данными о студентах
    Функция возвращает этот самый список

    filename - путь к excel-файлу
    solder - флаг, указывающий принадлежность таблицы
             (с данными о солдатах или об офицерах)
    """
    students_inform = []
    column_names = [
        "ФГОО ВО в котором обучается студент",  # 1
        "ФГОО ВО при которой создан ВУЦ",  # 2
        "ВУС",  # 3
        "ОВУ",  # 4
        "ФГОС",  # 5
        "Программа военной подготовки",  # 6
        "Год зачисления в ВУЗ",  # 7
        "Год начала обучения в ВУЦ",  # 8
        "Месяц нала обучения в ВУЦ",  # 9
        "Срок обучения (месяцев)",  # 10
        "Год окончания ВУЦ",  # 11
        "Месяц окончания обучения в ВУЦ",  # 12
        "Отметка о завершении военной подготовки",  # 13
        "Год окончания ВУЗа",  # 14
        "Месяц окончания обучения в ВУЗе",  # 15
        "Личный номер",  # 16
        "Фамилия",  # 17
        "Имя",  # 18
        "Отчество",  # 19
        "Дата рождения",  # 20
        "Место рождения",  # 21
        "Национальность",  # 22
        "Пол",  # 23
        "СНИЛС",  # 24
        "ИНН",  # 25
        "серия1",  # 26
        "серия2",  # 27
        "Номер",  # 28
        "Выдан",  # 29
        "Код подразделения",  # 30
        "Кем выдан",  # 31
        "Адрес регистрации",  # 32
        "Номер телефона",  # 33
        "Семейное положение",  # 34
        "Количество детей",  # 35
        "Военный комиссариат (по месту жительства)",  # 36
        "Военный комиссариат (в который будет направлено личное дело)",  # 37
        "По призыву",  # 38
        "По контракту",  # 39
        "Примечание",  # 40
        "Статус",  # 41
        "Приказ о зачислении",  # 42
        "Приказ о отчислении",  # 43
        "Причина отчисления"  # 44
    ]
    if solder:
        column_names.insert(2, "Код должности")
    wb = op.load_workbook(filename, data_only=True)
    sheet = wb.active
    bar = IncrementalBar(" ", max=sheet.max_row - 4)
    for student in range(5, sheet.max_row + 1):
        """
        Создаем из списка имён столбцов column_names и
        списка значений студента student_data словарик и
        добавляем каждый словарик в список students_inform
        """
        student_data = []
        for data in range(1, sheet.max_column + 1):
            student_data.append(sheet.cell(row=student, column=data).value)
        tmp = dict(zip(column_names, student_data))
        if solder and str(tmp["Код должности"]) != "256":
            tmp_vuc = tmp["Код должности"]
            tmp["Код должности"] = "256"
            tmp["ВУС"] = tmp_vuc
        students_inform.append(tmp)
        bar.next()
    return students_inform


def read_excel(filename: str) -> list:
    """
    Функция формирует список словарей с данными о студентах
    Функция возвращает этот самый список

    filename - путь к excel-файлу
    """
    wb = op.load_workbook(filename, data_only=True)
    sheet = wb.active
    if sheet.cell(row=1, column=4).value == "Код должности":
        return read_cicle(filename, True)
    if sheet.cell(row=1, column=4).value == "ОВУ":
        return read_cicle(filename, False)
    students_inform = []
    column_names = []
    for i in range(1, sheet.max_column + 1):
        column_names.append(sheet.cell(row=1, column=i).value)
    for student in trange(2, sheet.max_row + 1):
        """
        Создаем из списка имён столбцов column_names и
        списка значений студента student_data словарик и
        добавляем каждый словарик в список students_inform
        """
        student_data = []
        for data in range(1, sheet.max_column + 1):
            student_data.append(sheet.cell(row=student, column=data).value)
        tmp = dict(zip(column_names, student_data))
        for elem in tmp.values():
            if str(elem)[:5] == "солда" or str(elem)[:5] == "рядов":
                if ("Код должности" in tmp.keys()
                        and "ВУС" in tmp.keys() != "256"):
                    tmp_vuc = tmp["Код должности"]
                    tmp["Код должности"] = "256"
                    tmp["ВУС"] = tmp_vuc
        students_inform.append(tmp)
    return students_inform


def sum(list_one: list, list_two: list) -> list:
    """
    Функция сложения двух списков из двух эксель-таблиц
    Функция возвращает результирующий список

    list_one - первый список
    list_two - второй список
    """
    list_one = uniq_list(list_one)
    list_two = uniq_list(list_two)
    bar = IncrementalBar(" ", max=len(list_one))

    tmp = list_two
    for one in list_one:
        for two in list_two:
            if (str(one.get("Фамилия")).lower().strip()
                    == str(two.get("Фамилия")).lower().strip()):
                if (str(one.get("Отчество")).lower().strip()
                        == str(two.get("Отчество")).lower().strip()):
                    if (str(one.get("Имя")).lower().strip()
                            == str(two.get("Имя")).lower().strip()):
                        value = set(two) - set(one)
                        for new_key in list(value):
                            one[new_key] = two.get(new_key)
                        tmp.remove(two)
        bar.next()
    return list_one+tmp


def uniq_list(information: list) -> list:
    """
    Функция отбора уникальных параметров для одного человека
    Функция возвращает список словарей

    information - список с "разбросанными" параметрами
    """
    result = []
    for elem in information:
        tmp = {key: val for key, val in elem.items() if val and val != ""}
        if (len(tmp)) != 0:
            result.append(tmp)
    index_dublicate = []
    for i in range(len(result)-1):
        for j in range(i+1, len(result)-2):
            if (str(result[i].get("Имя")).lower().strip()
                    == str(result[j].get("Имя")).lower().strip()):
                if (str(result[i].get("Фамилия")).lower().strip()
                        == str(result[j].get("Фамилия")).lower().strip()):
                    if (str(result[i].get("Отчество")).lower().strip()
                            == str(result[j].get("Отчество")).lower().strip()):
                        if len(result[i]) <= len(result[j]):
                            value = set(result[j]) - set(result[i])
                            for new_key in list(value):
                                result[i][new_key] = result[j].get(new_key)
                        index_dublicate.append(j)
    index_dublicate = sorted(index_dublicate)
    for i in range(len(index_dublicate), 0):
        result.pop(index_dublicate[i]-i)
    return result


def write_excel(information: list, number: int) -> None:
    """
    Функция записывает данные в excel-таблицу

    information - список словарей
    number - номер опции
    """
    if number == 1:
        name_column = [
            "ФГОО ВО в котором обучается студент",  # 1
            "ФГОО ВО при которой создан ВУЦ",  # 2
            "ВУС",  # 3
            "ОВУ",  # 4
            "ФГОС",  # 5
            "Программа военной подготовки",  # 6
            "Год зачисления в ВУЗ",  # 7
            "Год начала обучения в ВУЦ",  # 8
            "Месяц нала обучения в ВУЦ",  # 9
            "Срок обучения (месяцев)",  # 10
            "Год окончания ВУЦ",  # 11
            "Месяц окончания обучения в ВУЦ",  # 12
            "Отметка о завершении военной подготовки",  # 13
            "Год окончания ВУЗа",  # 14
            "Месяц окончания обучения в ВУЗе",  # 15
            "Личный номер",  # 16
            "Фамилия",  # 17
            "Имя",  # 18
            "Отчество",  # 19
            "Дата рождения",  # 20
            "Место рождения",  # 21
            "Национальность",  # 22
            "Пол",  # 23
            "СНИЛС",  # 24
            "ИНН",  # 25
            "серия1",  # 26
            "серия2",  # 27
            "Номер",  # 28
            "Выдан",  # 29
            "Код подразделения",  # 30
            "Кем выдан",  # 31
            "Адрес регистрации",  # 32
            "Номер телефона",  # 33
            "Семейное положение",  # 34
            "Количество детей",  # 35
            "Военный комиссариат (по месту жительства)",  # 36
            "Военный комиссариат"
            " (в который будет направлено личное дело)",  # 37
            "По призыву",  # 38
            "По контракту",  # 39
            "Примечание",  # 40
            "Статус",  # 41
            "Приказ о зачислении",  # 42
            "Приказ о отчислении",  # 43
            "Причина отчисления"  # 44
        ]
        shutil.copyfile(
            "workprogram/"+"Не трогать!(Шаблон офицеров).xlsx",
            "Итоговые таблицы/(Итог)Офицеры.xlsx"
        )
        wb = op.load_workbook(
            "Итоговые таблицы/(Итог)Офицеры.xlsx",
            data_only=True
        )
        sheet = wb.active
        id = 0
        for j in range(len(information)):
            vuc = str(information[j].get("ВУС"))
            training = str(information[j].get("Программа военной подготовки"))
            if (training[:2] == "оф" or vuc == "461000"
                    or vuc == "461100" or vuc == "461200" or vuc == "461300"):
                for i in range(1, len(name_column)+1):
                    tmp = ""
                    if name_column[i-1] in information[j].keys():
                        tmp = information[j][name_column[i-1]]
                    sheet.cell(row=id+5, column=i, value=tmp)
                id += 1
        wb.save("Итоговые таблицы/(Итог)Офицеры.xlsx")
    elif number == 2:
        name_column = [
            "ФГОО ВО в котором обучается студент",  # 1
            "ФГОО ВО при которой создан ВУЦ",  # 2
            "ВУС",  # 3
            "Код должности",
            "ОВУ",  # 4
            "ФГОС",  # 5
            "Программа военной подготовки",  # 6
            "Год зачисления в ВУЗ",  # 7
            "Год начала обучения в ВУЦ",  # 8
            "Месяц нала обучения в ВУЦ",  # 9
            "Срок обучения (месяцев)",  # 10
            "Год окончания ВУЦ",  # 11
            "Месяц окончания обучения в ВУЦ",  # 12
            "Отметка о завершении военной подготовки",  # 13
            "Год окончания ВУЗа",  # 14
            "Месяц окончания обучения в ВУЗе",  # 15
            "Личный номер",  # 16
            "Фамилия",  # 17
            "Имя",  # 18
            "Отчество",  # 19
            "Дата рождения",  # 20
            "Место рождения",  # 21
            "Национальность",  # 22
            "Пол",  # 23
            "СНИЛС",  # 24
            "ИНН",  # 25
            "серия1",  # 26
            "серия2",  # 27
            "Номер",  # 28
            "Выдан",  # 29
            "Код подразделения",  # 30
            "Кем выдан",  # 31
            "Адрес регистрации",  # 32
            "Номер телефона",  # 33
            "Семейное положение",  # 34
            "Количество детей",  # 35
            "Военный комиссариат (по месту жительства)",  # 36
            "Военный комиссариат"
            " (в который будет направлено личное дело)",  # 37
            "По призыву",  # 38
            "По контракту",  # 39
            "Примечание",  # 40
            "Статус",  # 41
            "Приказ о зачислении",  # 42
            "Приказ о отчислении",  # 43
            "Причина отчисления"  # 44
        ]
        shutil.copyfile(
            "workprogram/"+"Не трогать!(Шаблон солдат).xlsx",
            "Итоговые таблицы/(Итог)Солдаты.xlsx"
        )
        wb = op.load_workbook(
            "Итоговые таблицы/(Итог)Солдаты.xlsx",
            data_only=True
        )
        sheet = wb.active
        id = 0
        for j in range(len(information)):
            vuc = str(information[j].get("ВУС"))
            training = str(information[j].get("Программа военной подготовки"))
            if (training[:2] == "со" or training[:2] == "ря" or vuc == "220"
                    or vuc == "233" or vuc == "250" or vuc == "262"):
                for i in range(1, len(name_column)+1):
                    tmp = ""
                    if name_column[i-1] in information[j].keys():
                        tmp = information[j][name_column[i-1]]
                    sheet.cell(row=id+5, column=i, value=tmp)
                id += 1
        wb.save("Итоговые таблицы/(Итог)Солдаты.xlsx")
    elif number == 3:
        print("своя таблица")
    else:
        name_column = []
        for elem in information:
            name_column = name_column + list(elem.keys())
        name_column = list(set(name_column))
        name_column.remove("Фамилия")
        name_column.remove("Имя")
        name_column.remove("Отчество")
        name_column.insert(0, "Фамилия")
        name_column.insert(1, "Имя")
        name_column.insert(2, "Отчество")
        wb = Workbook()
        ws = wb.active
        for i in range(1, len(name_column)+1):
            """
            Код, который выводит в первый столбец инфу с первого массива.
            Нужно сделать +- то же самое для остальных и все, по сути
            """
            ws.cell(row=1, column=i, value=name_column[i-1])
        for j in range(len(information)):
            for i in range(1, len(name_column)+1):
                tmp = ""
                if name_column[i-1] in information[j].keys():
                    tmp = information[j][name_column[i-1]]
                ws.cell(row=j+2, column=i, value=tmp)
        wb.save("Итоговые таблицы/Все данные.xlsx")


if __name__ == "__main__":
    all_files = listdir(getcwd()+"/Таблицы откуда берём информацию")
    excel_name = []
    for elem in all_files:
        filename, file_extension = path.splitext(elem)
        if file_extension == ".xlsx":
            excel_name.append(elem)
    print("Программа увидела следующие таблицы:")
    for elem in excel_name:
        print(elem)
    choose = input(
        "-------------------------------------------------"
        "-----------------------------\nПродожить? \n1)Да "
    )
    if choose == "1":
        information_list = []
        for elem in excel_name:
            print("\nЧитаем файл   "+str(elem))
            information_list.append(
                read_excel("Таблицы откуда берём информацию/"+elem)
            )
        print("\n" * 100)
        print("Данные успешно собраны")
        print("Совмещаем данные")
        if len(information_list) == 1:
            tmp = uniq_list(information_list[0])
            information_list.append(tmp)
            print(len(information_list))
            information_list.pop(0)
        while len(information_list) > 1:
            tmp = sum(information_list[0], information_list[1])
            information_list.pop(1)
            information_list.pop(0)
            information_list.append(tmp)
        print("\n" * 100)
        print("Данные совмещены")
        while True:
            tmp = input(
                "Какую таблицу создаём?\n1)Две таблицы: Солдаты, Офицеры\n"
                "2)Выгрузить все данные в произвольную таблицу "
            )
            if tmp == "1":
                write_excel(information_list[0], 1)
                write_excel(information_list[0], 2)
                input(
                    "Успешно, создали '(Итог)Офицеры.xlsx' и "
                    "'(Итог)Солдаты.xlsx'  в папке 'Итоговые таблицы'"
                )
            elif tmp == '2':
                write_excel(information_list[0], 4)
                input(
                    "Успешно,  создали 'Все данные.xlsx' "
                    "в папке 'Итоговые таблицы'"
                )
            else:
                input("Нет такого варианта...")
